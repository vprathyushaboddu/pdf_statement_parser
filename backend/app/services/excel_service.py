from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.models.transaction import Transaction

_COLUMNS = ["Date", "Description", "Debit", "Credit", "Balance"]


def build_workbook(transactions: list[Transaction]) -> Workbook:
    """Build the fixed-schema workbook described in requirements.md FR-13.

    Writes real date/number types (not strings) so totals/formulas work
    directly on the output (FR-10, FR-15).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws.append(_COLUMNS)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    for txn in sorted(transactions, key=lambda t: t.row_number):
        ws.append(
            [
                txn.txn_date,
                txn.description,
                float(txn.debit) if txn.debit is not None else None,
                float(txn.credit) if txn.credit is not None else None,
                float(txn.balance) if txn.balance is not None else None,
            ]
        )

    date_column = _COLUMNS.index("Date") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=date_column).number_format = "DD-MMM-YYYY"

    for idx, header in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, len(header) + 4)
    ws.column_dimensions[get_column_letter(_COLUMNS.index("Description") + 1)].width = 50

    return wb


def save_workbook(transactions: list[Transaction], destination_path: str) -> None:
    workbook = build_workbook(transactions)
    workbook.save(destination_path)
